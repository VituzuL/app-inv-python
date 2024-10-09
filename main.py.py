import os
import openpyxl
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.core.window import Window
from collections import defaultdict

class InventarioApp(App):

    def build(self):
        self.base_path = os.path.dirname(os.path.realpath(__file__))
        self.base_file = os.path.join(self.base_path, 'base.xlsx')
        self.load_base_data()

        # Dicionário para armazenar registros contados
        self.contagens = defaultdict(lambda: defaultdict(int))
        self.historico = []  # Histórico de lançamentos para correção

        # Interface gráfica construída diretamente no Python
        root = BoxLayout(orientation='vertical', padding=20, spacing=10)

        # Cabeçalho de seleção de contagem e tipo de inventário
        header_layout = BoxLayout(size_hint_y=None, height="40dp", spacing=10)
        header_layout.add_widget(Label(text="Contagem:", size_hint_x=None, width="100dp"))
        self.contagem_spinner = Spinner(text="Primeira", values=["Primeira", "Segunda", "Terceira"], size_hint_x=None, width="200dp")
        header_layout.add_widget(self.contagem_spinner)
        header_layout.add_widget(Label(text="Tipo:", size_hint_x=None, width="100dp"))
        self.tipo_spinner = Spinner(text="Insumo", values=["Insumo", "Embalagem"], size_hint_x=None, width="200dp")
        header_layout.add_widget(self.tipo_spinner)
        root.add_widget(header_layout)

        # Entrada de código de produto
        input_layout = BoxLayout(size_hint_y=None, height="40dp", spacing=10)
        input_layout.add_widget(Label(text="Código:", size_hint_x=None, width="80dp"))
        self.codigo_input = TextInput(multiline=False, size_hint_x=0.5)
        self.codigo_input.bind(on_text_validate=self.on_codigo_enter)
        search_button = Button(text="Buscar", size_hint_x=None, width="100dp")
        search_button.bind(on_press=self.on_search_button_press)
        input_layout.add_widget(self.codigo_input)
        input_layout.add_widget(search_button)
        root.add_widget(input_layout)

        # Exibição de descrição do produto
        self.descricao_label = Label(text="", size_hint_y=None, height="30dp", markup=True)
        root.add_widget(self.descricao_label)

        # Seleção de lote e entrada de quantidade
        lote_layout = BoxLayout(size_hint_y=None, height="40dp", spacing=10)
        lote_layout.add_widget(Label(text="Lote:", size_hint_x=None, width="80dp"))
        self.lotes_spinner = Spinner(text="Selecione um lote", values=[], size_hint_x=None, width="200dp")
        lote_layout.add_widget(self.lotes_spinner)
        self.novo_lote_input = TextInput(multiline=False, hint_text="Novo Lote", size_hint_x=None, width="200dp")
        lote_layout.add_widget(self.novo_lote_input)
        root.add_widget(lote_layout)

        quantidade_layout = BoxLayout(size_hint_y=None, height="40dp", spacing=10)
        quantidade_layout.add_widget(Label(text="Quantidade:", size_hint_x=None, width="100dp"))
        self.quantidade_input = TextInput(multiline=False, hint_text="Quantidade", input_filter='int', size_hint_x=None, width="100dp")
        quantidade_layout.add_widget(self.quantidade_input)
        root.add_widget(quantidade_layout)

        # Botões para salvar contagem, limpar lançamentos e corrigir último lançamento
        button_layout = BoxLayout(size_hint_y=None, height="40dp", spacing=10)
        save_button = Button(text="Salvar Contagem")
        save_button.bind(on_press=self.salvar_contagem)
        button_layout.add_widget(save_button)
        limpar_button = Button(text="Limpar Lançamentos")
        limpar_button.bind(on_press=self.limpar_lancamentos)
        button_layout.add_widget(limpar_button)
        corrigir_button = Button(text="Corrigir Último Lançamento")
        corrigir_button.bind(on_press=self.corrigir_ultimo_lancamento)
        button_layout.add_widget(corrigir_button)
        export_button = Button(text="Exportar para Excel")
        export_button.bind(on_press=self.exportar_para_excel)
        button_layout.add_widget(export_button)
        root.add_widget(button_layout)

        # Exibição dos registros contados
        self.registros_layout = GridLayout(cols=1, size_hint_y=None)
        self.registros_layout.bind(minimum_height=self.registros_layout.setter('height'))
        scroll_view = ScrollView(size_hint=(1, None), size=(Window.width, Window.height/2))
        scroll_view.add_widget(self.registros_layout)
        root.add_widget(scroll_view)

        return root

    def load_base_data(self):
        # Carregar dados do arquivo Excel
        self.produtos = self.load_produtos()
        self.estoque = self.load_estoque()

    def load_produtos(self):
        wb = openpyxl.load_workbook(self.base_file)
        produtos_sheet = wb['Produtos']
        produtos = {}
        for row in produtos_sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, tipo = row[0], row[1], row[2]
            produtos[str(codigo).strip()] = {'descricao': descricao, 'tipo': tipo, 'lotes': []}
        wb.close()
        return produtos

    def load_estoque(self):
        wb = openpyxl.load_workbook(self.base_file)
        estoque_sheet = wb['Estoque']
        for row in estoque_sheet.iter_rows(min_row=2, values_only=True):
            codigo, descricao, lote = row[0], row[1], row[2]
            if str(codigo).strip() in self.produtos:
                self.produtos[str(codigo).strip()]['lotes'].append(lote)
        wb.close()

    def on_codigo_enter(self, instance):
        codigo = instance.text.strip()
        self.update_product_info(codigo)

    def on_search_button_press(self, instance):
        codigo = self.codigo_input.text.strip()
        self.update_product_info(codigo)

    def update_product_info(self, codigo):
        if codigo in self.produtos:
            self.descricao_label.text = f"[b]Descrição:[/b] {self.produtos[codigo]['descricao']}"
            self.lotes_spinner.values = self.produtos[codigo]['lotes']
            self.lotes_spinner.text = "Selecione um lote"
            self.novo_lote_input.text = ""
            self.quantidade_input.text = ""
        else:
            self.descricao_label.text = '[b]Produto não encontrado[/b]'
            self.lotes_spinner.values = []
            self.lotes_spinner.text = "Selecione um lote"
            self.novo_lote_input.text = ""
            self.quantidade_input.text = ""

    def salvar_contagem(self, instance):
        codigo = self.codigo_input.text.strip()
        if not codigo or codigo not in self.produtos:
            self.show_popup("Erro", "Código do produto inválido ou não encontrado.")
            return

        descricao = self.produtos[codigo]['descricao']
        lote = self.lotes_spinner.text if self.lotes_spinner.text != "Selecione um lote" else self.novo_lote_input.text.strip()
        quantidade = self.quantidade_input.text.strip()

        if not lote:
            self.show_popup("Erro", "Lote não pode estar vazio.")
            return

        if not quantidade:
            self.show_popup("Erro", "Quantidade não pode estar vazia.")
            return

        quantidade = int(quantidade)
        # Atualiza o dicionário de contagens
        self.contagens[codigo][lote] += quantidade
        # Salvar no histórico para possível correção
        self.historico.append((codigo, lote, quantidade))

        # Atualiza a exibição dos registros contados
        self.update_registros_layout()

        # Limpa os campos de entrada
        self.codigo_input.text = ""
        self.descricao_label.text = ""
        self.lotes_spinner.text = "Selecione um lote"
        self.lotes_spinner.values = []
        self.novo_lote_input.text = ""
        self.quantidade_input.text = ""

    def update_registros_layout(self):
        # Limpa o layout atual
        self.registros_layout.clear_widgets()
        # Adiciona novamente todos os registros contados
        for codigo, lotes in self.contagens.items():
            for lote, quantidade in lotes.items():
                descricao = self.produtos[codigo]['descricao']
                registro_label = Label(text=f"Código: {codigo} | Descrição: {descricao} | Lote: {lote} | Quantidade: {quantidade}",
                                       size_hint_y=None, height="30dp")
                self.registros_layout.add_widget(registro_label)

    def limpar_lancamentos(self, instance):
        # Limpar o dicionário de contagens e o histórico
        self.contagens.clear()
        self.historico.clear()
        # Atualizar a exibição dos registros
        self.update_registros_layout()

    def corrigir_ultimo_lancamento(self, instance):
        if self.historico:
            codigo, lote, quantidade = self.historico.pop()
            if self.contagens[codigo][lote] <= quantidade:
                del self.contagens[codigo][lote]
            else:
                self.contagens[codigo][lote] -= quantidade
            # Atualizar a exibição dos registros
            self.update_registros_layout()
        else:
            self.show_popup("Erro", "Não há lançamentos para corrigir.")

    def exportar_para_excel(self, instance):
        contagem = self.contagem_spinner.text
        tipo = self.tipo_spinner.text
        filename = f"{contagem.lower()}_contagem_{tipo.lower()}.xlsx"
        filepath = os.path.join(self.base_path, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contagem"

        # Cabeçalhos fixos
        ws.append(["COD_EMPRESA", "COD_FILIAL", "COD_DEPOSITO", "Código", "Descrição", "Lote", "Quantidade"])
        cod_empresa = "1010"
        cod_filial = "M016"
        cod_deposito = "GM01"

        # Adiciona as contagens ao arquivo Excel
        for codigo, lotes in self.contagens.items():
            descricao = self.produtos[codigo]['descricao']
            for lote, quantidade in lotes.items():
                ws.append([cod_empresa, cod_filial, cod_deposito, codigo, descricao, lote, quantidade])

        wb.save(filepath)
        self.show_popup("Sucesso", f"Arquivo exportado como {filename}")

    def show_popup(self, title, message):
        popup_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        popup_layout.add_widget(Label(text=message))
        close_button = Button(text="Fechar", size_hint_y=None, height="40dp")
        popup_layout.add_widget(close_button)

        popup = Popup(title=title, content=popup_layout, size_hint=(0.8, 0.4))
        close_button.bind(on_press=popup.dismiss)
        popup.open()

if __name__ == '__main__':
    InventarioApp().run()
